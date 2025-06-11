from flask import Flask, request, jsonify
import requests
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

OPENROUTER_API_KEY = "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"
MODEL = "openai/gpt-3.5-turbo"
EXCEL_FILE = "chat_log.xlsx"

SYSTEM_PROMPT = (
    "You are the official AI assistant of Simko. "
    "Simko does not have an investor yet. "
    "Your role is to explain Simko's vision, approach, and innovations in a clear, friendly, and technically accurate way. "
    "Simko is redefining electric vehicles by reducing weight, cost, and environmental impact through custom motors, compact drivetrains, and lightweight design. "
    "Help visitors understand how Simko achieves 30–50% EV weight savings, and supports a carbon-negative future. "
    "Always answer all questions the user asks, even if they include multiple questions in a single message."
)

def log_to_excel(message, response):
    if not os.path.isfile(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Chat Log"
        ws.append(["Timestamp", "User Question", "Bot Response"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), message, response])
    wb.save(EXCEL_FILE)

@app.route("/chat", methods=["POST"])
def chat():
    data = request.get_json()
    user_input = data.get("message", "")

    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_input}
        ]
    }

    try:
        response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload, timeout=10)
    except Exception as e:
        return jsonify({"response": f"Failed to connect to OpenRouter API: {str(e)}"}), 500

    if response.status_code == 200:
        reply = response.json()["choices"][0]["message"]["content"]
        log_to_excel(user_input, reply)
        return jsonify({"response": reply})
    else:
        print("Status code:", response.status_code)
        print("Error text:", response.text)
        return jsonify({"response": f"Error: {response.status_code}, {response.text}"}), response.status_code

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)
